from flask import Flask, request, jsonify, session, redirect, url_for, send_file
from flask_cors import CORS
from flask_sqlalchemy import SQLAlchemy
from werkzeug.utils import secure_filename
import pandas as pd
import os
import re
import requests
import json
from datetime import datetime, timedelta
import logging
from apscheduler.schedulers.background import BackgroundScheduler
import atexit
from functools import wraps
import pytz
from config import config
import threading
import uuid

# 配置日志
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)

# 根据环境变量选择配置
config_name = os.environ.get('FLASK_ENV', 'development')
app.config.from_object(config[config_name])

# 时区配置
TIMEZONE = pytz.timezone('Asia/Shanghai')

def get_local_time():
    """获取本地时间（Asia/Shanghai时区）"""
    return datetime.now(TIMEZONE).replace(tzinfo=None)

def format_datetime(dt):
    """格式化日期时间为字符串"""
    if dt is None:
        return None
    return dt.strftime('%Y-%m-%d %H:%M:%S')

# 有权限的用户学工号列表（从配置文件读取，不再硬编码）
# 授权用户列表已在 config.py 中配置，这里不再覆盖
# app.config['AUTHORIZED_USERS'] 已经在 Config 类中设置

# 初始化扩展
db = SQLAlchemy(app)
CORS(app, supports_credentials=True)

# 任务管理：存储异步任务的进度信息
# 格式：{task_id: {status, total, processed, success, failed, start_time, end_time, failed_students}}
task_progress = {}
task_lock = threading.Lock()

def clean_old_tasks():
    """清理超过1小时的已完成任务"""
    with task_lock:
        current_time = get_local_time()
        tasks_to_remove = []
        
        for task_id, progress in task_progress.items():
            if progress['status'] == 'completed' and progress['end_time']:
                end_time = datetime.fromisoformat(progress['end_time'])
                if (current_time - end_time).total_seconds() > 3600:  # 1小时
                    tasks_to_remove.append(task_id)
        
        for task_id in tasks_to_remove:
            del task_progress[task_id]
            logger.info(f"清理过期任务: {task_id}")

# 确保上传目录存在
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# 数据库模型
class LawSchoolUploadList(db.Model):
    __tablename__ = 'lawSchoolUploadList'
    
    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.String(50), nullable=False, comment='学号')
    name = db.Column(db.String(100), nullable=False, comment='姓名')
    gender = db.Column(db.String(10), comment='性别')
    nationality = db.Column(db.String(50), comment='民族')
    political_status = db.Column(db.String(50), comment='政治面貌')
    id_type = db.Column(db.String(50), comment='证件类型')
    id_number = db.Column(db.String(50), comment='证件号码')
    country = db.Column(db.String(50), comment='国家地区')
    hong_kong_macao_taiwan = db.Column(db.String(50), comment='港澳台')
    campus_email = db.Column(db.String(100), comment='校内邮箱')
    personal_email = db.Column(db.String(100), comment='个人邮箱')
    phone = db.Column(db.String(20), comment='手机号码')
    major = db.Column(db.String(100), comment='专业')
    department = db.Column(db.String(100), comment='院系')
    study_mode = db.Column(db.String(50), comment='学习方式')
    grade = db.Column(db.String(20), comment='年级')
    enrollment_date = db.Column(db.Date, comment='入学日期')
    student_type = db.Column(db.String(50), comment='学生类型')
    degree_type = db.Column(db.String(50), comment='学位类型')
    special_plan = db.Column(db.String(100), comment='专项计划')
    degree_level = db.Column(db.String(50), comment='学位层次')
    expected_graduation_date = db.Column(db.Date, comment='预计毕业时间')
    enrollment_mode = db.Column(db.String(50), comment='入学方式')
    student_category = db.Column(db.String(50), comment='学生类别')
    supervisor = db.Column(db.String(100), comment='导师姓名')
    class_name = db.Column(db.String(100), comment='班级')
    registration_status = db.Column(db.String(50), comment='注册状态')
    z308_id = db.Column(db.String(50), comment='Aleph系统ID，有值表示有aleph账号，NULL表示无aleph账号')
    sync_z308_id = db.Column(db.Integer, default=0, comment='同步状态：0-未同步，1-已同步，2-同步失败（仅当z308_id有值时才有意义）')
    uploader_id = db.Column(db.String(50), nullable=False, comment='上传人学工号')
    create_time = db.Column(db.DateTime, default=get_local_time, comment='创建时间')
    update_time = db.Column(db.DateTime, default=get_local_time, onupdate=get_local_time, comment='更新时间')


# 登录验证装饰器
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'success': False, 'message': '请先登录'}), 401
        return f(*args, **kwargs)
    return decorated_function

# 权限验证装饰器
def permission_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'success': False, 'message': '请先登录'}), 401
        
        # 检查用户是否在授权列表中
        if session['user_id'] not in app.config['AUTHORIZED_USERS']:
            return jsonify({'success': False, 'message': '用户无权限访问'}), 403
        return f(*args, **kwargs)
    return decorated_function

# 外部接口调用函数
def call_api(url, payload):
    """调用外部API"""
    try:
        import json
        logger.info(f"发送的POST请求到: {url}")
        logger.info(f"请求参数: {payload}")
        
        headers = {
            'Content-Type': 'application/json'
        }
        
        response = requests.post(
            url,
            json=payload,
            data=json.dumps(payload),
            timeout=30
        )
        response.raise_for_status()
        return response.json()
    except requests.exceptions.RequestException as e:
        logger.error(f"API调用失败: {e}")
        # 尝试获取响应内容以便调试
        try:
            if hasattr(e, 'response') and e.response is not None:
                logger.error(f"响应状态码: {e.response.status_code}")
                logger.error(f"响应内容: {e.response.text}")
        except:
            pass
        return None

def call_encrypt_api(qry_str, qry_type):
    """调用加解密接口"""
    logger.info(f"当前API_KEY配置值: {app.config['API_KEY']}")
    payload = {
        "userCode": app.config["API_USERCODE"],
        "userPwd": app.config["API_PWD"],
        "qryType": qry_type,  # 调用类型，1=加密，2=解密
        "qryKey": app.config["API_KEY"],
        "qryStr": qry_str
    }
    logger.info(f"加解密请求参数: {payload}")
    
    # 调用加解密接口
    encrypt_res = call_api(
        app.config["API_URL"] + "/getEncryptCode.ashx",
        payload
    )
    logger.info(f"加解密响应: {encrypt_res}")
    
    # 验证接口响应
    if not encrypt_res or encrypt_res.get('resStr') != '1':
        error_msg = encrypt_res.get('msgStr', '未知错误') if encrypt_res else '接口调用失败'
        logger.error(f"加解密失败: {error_msg}")
        return {
            'success': False,
            'message': f"加解密失败: {error_msg}"
        }
    
    # 解析返回结果，提取encodeStr字段
    encrypt_data = {
        'user_code': encrypt_res.get('userCode'),
        'qry_key': encrypt_res.get('qryKey'),  
        'qry_str': encrypt_res.get('qryStr'),
        'encode_str': encrypt_res.get('encodeStr')  # 重点解析的字段
    }
    
    return {
        'success': True,
        'data': encrypt_data['encode_str'],
        'message': encrypt_res.get('msgStr', '加解密成功')
    }

def call_z308_id_api(student_id):
    """调用学工号查询z308_id接口"""
    payload = {
        "userCode": app.config["API_USERCODE"],
        "userPwd": app.config["API_PWD"],
        "qryType": "1",  # 固定值1
        "qryStr": student_id  # 学工号
    }
    
    response = call_api(
        app.config["API_URL"] + "/getAlephSysID.ashx",
        payload
    )
    
    if not response or response.get('resStr') != '1':
        return None
    
    ret_user = response.get('retUser', [])
    if ret_user and len(ret_user) > 0:
        return ret_user[0].get('sysId')
    
    return None

def call_sync_z308_id_api(z308_id, student_data):
    """
    调用同步z308_id接口，为学生在Aleph系统中添加法学院分馆权限
    :param z308_id: Aleph系统ID
    :param student_data: 学生数据字典
    :return: (success: bool, message: str) 成功返回(True, '授权成功')，失败返回(False, '错误信息')
    """
    try:
        # 构造请求参数
        payload = {
            "userCode": app.config["API_USERCODE"],
            "userPwd": app.config["API_PWD"],
            "z308Id": z308_id,  # 注意：接口要求驼峰命名的 z308Id
            "sublib": "FAL"     # 法学院分馆代码
        }
        
        logger.info(f"调用同步接口，z308_id: {z308_id}, 学号: {student_data.get('student_id')}")
        logger.info(f"请求参数: {payload}")
        
        # 调用外部API
        response = call_api(
            "http://10.119.4.239/docaffiresinterface/addPrevByZ308ID.ashx",
            payload
        )
        
        logger.info(f"同步接口返回: {response}")
        
        if response:
            ret_state = response.get('retState')
            ret_msg = response.get('retMsg', '未知错误')
            
            # retState = 1 或 5 都认为是成功
            # 1: 授权成功
            # 5: 授权失败，原因是：已经有该分馆权限（这种情况也算成功）
            if ret_state == 1:
                logger.info(f"同步成功: {ret_msg}")
                return (True, '授权成功')
            elif ret_state == 5:
                logger.info(f"用户已有权限: {ret_msg}")
                return (True, '已有分馆权限')
            else:
                # 其他情况都是失败
                # -3: 用户认证失败
                # -4: 发生不可预期的错误
                # 3: z308ID格式不正确
                # 4: 此z308Id有误，尚未有主馆权限
                logger.error(f"同步失败: retState={ret_state}, retMsg={ret_msg}")
                return (False, ret_msg)
        else:
            logger.error("同步接口无响应")
            return (False, '同步接口无响应')
            
    except Exception as e:
        logger.error(f"调用同步接口异常: {e}")
        return (False, f'调用同步接口异常: {str(e)}')

# 异步任务：批量查询 z308_id
def async_query_z308_ids(task_id, student_records):
    """
    异步查询学生的 z308_id 并同步授权
    :param task_id: 任务ID
    :param student_records: 学生记录列表 [(record_id, student_id, name), ...]
    """
    # 在异步线程中需要创建应用上下文
    with app.app_context():
        with task_lock:
            task_progress[task_id] = {
                'status': 'processing',
                'total': len(student_records),
                'processed': 0,
                'success': 0,
                'failed': 0,
                'start_time': get_local_time().isoformat(),
                'end_time': None,
                'failed_students': []
            }
        
        logger.info(f"任务 {task_id} 开始：共 {len(student_records)} 个学生需要查询 z308_id 并同步授权")
        
        for record_id, student_id, name in student_records:
            try:
                # 步骤1：查询 z308_id
                z308_id = call_z308_id_api(student_id)
                
                # 更新数据库
                record = LawSchoolUploadList.query.get(record_id)
                if record:
                    record.z308_id = z308_id
                    
                    if z308_id:
                        logger.info(f"学生 {student_id} 查询到 z308_id: {z308_id}，开始同步授权...")
                        
                        # 步骤2：如果有z308_id，调用同步接口进行授权
                        student_data = {
                            'student_id': record.student_id,
                            'name': record.name,
                            'gender': record.gender,
                            'nationality': record.nationality,
                            'political_status': record.political_status,
                            'id_type': record.id_type,
                            'id_number': record.id_number,
                            'country': record.country,
                            'hong_kong_macao_taiwan': record.hong_kong_macao_taiwan,
                            'campus_email': record.campus_email,
                            'personal_email': record.personal_email,
                            'phone': record.phone,
                            'major': record.major,
                            'department': record.department,
                            'study_mode': record.study_mode,
                            'grade': record.grade,
                            'enrollment_date': record.enrollment_date.strftime('%Y-%m-%d') if record.enrollment_date else None,
                            'student_type': record.student_type,
                            'degree_type': record.degree_type,
                            'special_plan': record.special_plan,
                            'degree_level': record.degree_level,
                            'expected_graduation_date': record.expected_graduation_date.strftime('%Y-%m-%d') if record.expected_graduation_date else None,
                            'enrollment_mode': record.enrollment_mode,
                            'student_category': record.student_category,
                            'supervisor': record.supervisor,
                            'class_name': record.class_name,
                            'registration_status': record.registration_status
                        }
                        
                        # 调用同步接口
                        sync_success, sync_message = call_sync_z308_id_api(z308_id, student_data)
                        
                        if sync_success:
                            # 同步成功
                            record.sync_z308_id = 1
                            logger.info(f"学生 {student_id} 同步成功: {sync_message}")
                            with task_lock:
                                task_progress[task_id]['success'] += 1
                        else:
                            # 同步失败
                            record.sync_z308_id = 2
                            logger.error(f"学生 {student_id} 同步失败: {sync_message}")
                            with task_lock:
                                task_progress[task_id]['failed'] += 1
                                task_progress[task_id]['failed_students'].append({
                                    'student_id': student_id,
                                    'name': name,
                                    'error': f'同步失败: {sync_message}'
                                })
                    else:
                        # 未查询到z308_id，说明没有aleph账号
                        logger.warning(f"学生 {student_id} 没有 Aleph 账号")
                        record.sync_z308_id = 0  # 保持为0，表示没有z308_id
                        with task_lock:
                            task_progress[task_id]['failed'] += 1
                            task_progress[task_id]['failed_students'].append({
                                'student_id': student_id,
                                'name': name,
                                'error': '没有Aleph账号'
                            })
                    
                    record.update_time = get_local_time()
                    db.session.commit()
                
            except Exception as e:
                logger.error(f"处理学生 {student_id} 失败: {e}")
                db.session.rollback()  # 回滚当前事务
                with task_lock:
                    task_progress[task_id]['failed'] += 1
                    task_progress[task_id]['failed_students'].append({
                        'student_id': student_id,
                        'name': name,
                        'error': str(e)
                    })
            
            # 更新进度
            with task_lock:
                task_progress[task_id]['processed'] += 1
        
        # 任务完成
        with task_lock:
            task_progress[task_id]['status'] = 'completed'
            task_progress[task_id]['end_time'] = get_local_time().isoformat()
        
        logger.info(f"任务 {task_id} 完成：成功 {task_progress[task_id]['success']}，失败 {task_progress[task_id]['failed']}")

# 辅助函数：清理手机号格式
def clean_phone_number(phone_value):
    """清理手机号，去除Excel导致的格式问题"""
    if not phone_value or pd.isna(phone_value):
        return None
    
    phone = str(phone_value).strip()
    if not phone or phone.lower() == 'nan':
        return None
    
    # 清理可能的.0后缀（Excel数字格式导致）
    if phone.endswith('.0'):
        phone = phone[:-2]
    
    # 移除所有空格、连字符等
    phone = phone.replace(' ', '').replace('-', '').replace('(', '').replace(')', '')
    
    return phone if phone else None

# 数据验证函数
def validate_student_data(row, row_number):
    """验证学生数据"""
    errors = []
    student_id = str(row.get('学号', '')).strip() if pd.notna(row.get('学号')) else ''
    name = str(row.get('姓名', '')).strip() if pd.notna(row.get('姓名')) else ''
    
    # 必填字段验证
    if not student_id:
        errors.append('学号不能为空')
    if not name:
        errors.append('姓名不能为空')
    # 证件号码改为非必填，不验证是否为空
    # 手机号码改为非必填，不验证是否为空
    
    # 学号格式验证（假设学号格式为数字或字母数字组合）
    if student_id and not re.match(r'^[A-Za-z0-9]+$', student_id):
        errors.append(f'学号格式不正确（只能包含字母和数字）')
    
    # 手机号格式验证（仅在填写时验证格式）
    phone = clean_phone_number(row.get('手机号码'))
    if phone:  # 只有填写了手机号才验证格式
        if not re.match(r'^1[3-9]\d{9}$', phone):
            errors.append(f'手机号格式不正确（应为11位数字，以1开头）')
    
    # 邮箱格式验证
    campus_email = str(row.get('校内电子邮箱', '')).strip() if pd.notna(row.get('校内电子邮箱')) else ''
    personal_email = str(row.get('电子邮箱', '')).strip() if pd.notna(row.get('电子邮箱')) else ''
    
    email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    
    if campus_email and not re.match(email_pattern, campus_email):
        errors.append(f'校内电子邮箱格式不正确')
    
    if personal_email and not re.match(email_pattern, personal_email):
        errors.append(f'电子邮箱格式不正确')
    
    # 证件号码格式验证（支持身份证、护照等多种证件类型）
    id_number = str(row.get('证件号码', '')).strip() if pd.notna(row.get('证件号码')) else ''
    if id_number:
        # 移除可能的空格
        id_number = id_number.replace(' ', '')
        
        # 证件号码长度应在5-30位之间（覆盖大部分证件类型）
        if len(id_number) < 5 or len(id_number) > 30:
            errors.append(f'证件号码长度不正确（应为5-30位）')
        # 证件号码只能包含字母、数字和常见符号
        elif not re.match(r'^[A-Za-z0-9\-]+$', id_number):
            errors.append(f'证件号码格式不正确（只能包含字母、数字和连字符）')
    
    # 如果有错误，返回带有学号和姓名的详细信息
    if errors:
        student_info = f"学号: {student_id}, 姓名: {name}" if student_id or name else "未填写学号和姓名"
        return [f"{student_info} - {error}" for error in errors]
    
    return []

# API路由
@app.route('/lawSchUpl/api/login/encrypt', methods=['POST', 'GET'])
def login_encrypt():
    """获取登录加密URL或处理登录回调"""
    try:
        # 如果是GET请求且有token参数，说明是登录回调
        if request.method == 'GET' and request.args.get('token'):
            token = request.args.get('token')
            logger.info(f"收到登录回调，token: {token}")
            
            # 去掉token末尾的sjtulibt
            if token.endswith('sjtulibt'):
                token = token[:-8]  # 去掉末尾的8个字符'sjtulibt'
            
            result = call_encrypt_api(token, '2')  # 2表示解密
            
            if result['success']:
                user_id = result['data']
                
                # 检查用户权限
                if user_id not in app.config['AUTHORIZED_USERS']:
                    return jsonify({'success': False, 'message': '用户无权限访问系统'})
                
                # 设置session
                session['user_id'] = user_id
                session['user_name'] = user_id  # 使用学工号作为显示名称
                
                # 返回JSON响应而不是重定向
                return jsonify({
                    'success': True,
                    'user_id': user_id,
                    'user_name': user_id,
                    'message': '登录成功'
                })
            else:
                return jsonify(result)
        
        # 如果是POST请求，处理加密请求
        elif request.method == 'POST':
            data = request.get_json() or {}
            # 从请求体中获取当前URL
            current_url = data.get('current_url', '')
            logger.info(f"从请求体获取的URL: {current_url}")
            
            if not current_url:
                # 如果请求体没有传递，尝试从请求头获取
                current_url = request.headers.get('X-Current-URL', '')
                logger.info(f"从X-Current-URL获取的URL: {current_url}")
                
            if not current_url:
                # 如果都没有，使用默认的登录成功跳转地址
                # 从请求头中获取Host，构建正确的前端URL
                host = request.headers.get('Host', 'localhost')
                current_url = f"http://{host}/lawSchUpl/login/callback"
                logger.info(f"使用默认URL: {current_url}")
            
            logger.info(f"最终使用的回调URL: {current_url}")
            
            if not current_url:
                return jsonify({'success': False, 'message': '当前URL不能为空'})
            
            result = call_encrypt_api(current_url, '1')  # 1表示加密
            
            if result['success']:
                login_url = f"http://10.119.4.239/docaffiresinterface/userIdentify.aspx?codeStr={result['data']}sjtulibt"
                return jsonify({
                    'success': True,
                    'loginUrl': login_url,
                    'message': result['message']
                })
            else:
                return jsonify(result)
        
        else:
            return jsonify({'success': False, 'message': '请求方法不支持'})
            
    except Exception as e:
        logger.error(f"登录处理失败: {e}")
        return jsonify({'success': False, 'message': '系统错误'}), 500

@app.route('/lawSchUpl/api/login/callback', methods=['GET'])
def login_callback():
    """登录回调处理（GET方式）"""
    try:
        token = request.args.get('token')
        if not token:
            return jsonify({'success': False, 'message': 'Token不能为空'})
        
        # 去掉token末尾的sjtulibt
        if token.endswith('sjtulibt'):
            token = token[:-8]  # 去掉末尾的8个字符'sjtulibt'
        
        result = call_encrypt_api(token, '2')  # 2表示解密
        
        if result['success']:
            user_id = result['data']
            
            # 检查用户权限
            if user_id not in app.config['AUTHORIZED_USERS']:
                return jsonify({'success': False, 'message': '用户无权限访问系统'})
            
            # 设置session
            session['user_id'] = user_id
            session['user_name'] = user_id  # 使用学工号作为显示名称
            
            # 返回JSON响应而不是重定向
            return jsonify({
                'success': True,
                'user_id': user_id,
                'user_name': user_id,
                'message': '登录成功'
            })
        else:
            return jsonify(result)
            
    except Exception as e:
        logger.error(f"登录解密失败: {e}")
        return jsonify({'success': False, 'message': f"登录解密失败: {e}"}), 500

@app.route('/api/login/decrypt', methods=['POST'])
def login_decrypt():
    """解密登录token获取用户学工号"""
    try:
        data = request.get_json()
        token = data.get('token', '')
        
        if not token:
            return jsonify({'success': False, 'message': 'Token不能为空'})
        
        result = call_encrypt_api(token, '2')  # 2表示解密
        
        if result['success']:
            user_id = result['data']
            
            # 检查用户权限
            if user_id not in app.config['AUTHORIZED_USERS']:
                return jsonify({'success': False, 'message': '用户无权限访问系统'})
            
            # 设置session
            session['user_id'] = user_id
            session['user_name'] = user_id  # 使用学工号作为显示名称
            
            # 重定向到前端页面
            return redirect('http://localhost/lawSchUpl/')
        else:
            return jsonify(result)
            
    except Exception as e:
        logger.error(f"登录解密失败: {e}")
        return jsonify({'success': False, 'message': '系统错误'}), 500

@app.route('/lawSchUpl/api/logout', methods=['POST'])
def logout():
    """退出登录"""
    session.clear()
    return jsonify({'success': True, 'message': '退出成功'})

@app.route('/lawSchUpl/api/user/info', methods=['GET'])
@login_required
def get_user_info():
    """获取当前用户信息"""
    return jsonify({
        'success': True,
        'user_id': session.get('user_id'),
        'user_name': session.get('user_name')
    })

@app.route('/lawSchUpl/api/upload/template', methods=['GET'])
@permission_required
def download_template():
    """下载Excel模板"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = '学生名单'
        
        # 定义列名
        columns = [
            '学号', '姓名', '性别', '民族', '政治面貌', '证件类型', '证件号码',
            '国家地区', '港澳台侨', '校内电子邮箱', '电子邮箱', '手机号码',
            '专业', '院系', '学习方式', '年级', '入学年月', '入学方式名称',
            '培养层次', '专项计划', '学位类型', '预计毕业时间', '培养方式名称',
            '学生类别', '导师姓名', '班级', '注册状态'
        ]
        
        # 第1行：添加重要提示信息
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
        warning_cell = ws.cell(row=1, column=1)
        warning_cell.value = '⚠️ 重要提示：下面第3行是示例数据，上传前请务必删除示例数据行，然后填入真实学生信息。必填字段：学号、姓名'
        warning_cell.font = Font(name='微软雅黑', size=12, bold=True, color='FF0000')  # 红色加粗
        warning_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        warning_cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # 黄色背景
        ws.row_dimensions[1].height = 40
        
        # 第2行：表头
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')  # 蓝色背景
        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')  # 白色加粗
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        for col_num, column_title in enumerate(columns, 1):
            cell = ws.cell(row=2, column=col_num)
            cell.value = column_title
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        ws.row_dimensions[2].height = 20
        
        # 第3行：示例数据（只有一行）
        example_data = [
            '025190210001',           # 学号
            '张三',                   # 姓名
            '男',                     # 性别
            '汉族',                   # 民族
            '中共党员',               # 政治面貌
            '身份证',                 # 证件类型
            '110101199001011234',     # 证件号码
            '中华人民共和国',         # 国家地区
            '',                       # 港澳台侨
            'zhangsan@sjtu.edu.cn',   # 校内电子邮箱
            'zhangsan@example.com',   # 电子邮箱
            '13900000000',            # 手机号码
            '法学(030100)',           # 专业
            '(190)凯原法学院',        # 院系
            '全日制',                 # 学习方式
            '2025级',                 # 年级
            '2025-09-01',             # 入学年月
            '入学申请制',             # 入学方式名称
            '博士',                   # 培养层次
            '',                       # 专项计划
            '学术学位',               # 学位类型
            '2029-06-30',             # 预计毕业时间
            '非定向',                 # 培养方式名称
            '学术型博士',             # 学生类别
            '李四',                   # 导师姓名
            '法学25D101',             # 班级
            '已注册'                  # 注册状态
        ]
        
        example_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')  # 浅黄色背景
        example_alignment = Alignment(horizontal='left', vertical='center')
        
        for col_num, value in enumerate(example_data, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.value = value
            cell.fill = example_fill
            cell.alignment = example_alignment
        
        ws.row_dimensions[3].height = 18
        
        # 设置列宽
        column_widths = [
            15,  # 学号
            10,  # 姓名
            8,   # 性别
            10,  # 民族
            15,  # 政治面貌
            10,  # 证件类型
            20,  # 证件号码
            18,  # 国家地区
            12,  # 港澳台侨
            25,  # 校内电子邮箱
            25,  # 电子邮箱
            13,  # 手机号码
            18,  # 专业
            18,  # 院系
            10,  # 学习方式
            10,  # 年级
            12,  # 入学年月
            15,  # 入学方式名称
            10,  # 培养层次
            15,  # 专项计划
            12,  # 学位类型
            15,  # 预计毕业时间
            15,  # 培养方式名称
            15,  # 学生类别
            12,  # 导师姓名
            15,  # 班级
            12   # 注册状态
        ]
        
        from openpyxl.utils import get_column_letter
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width
        
        # 保存Excel文件
        template_path = os.path.join(app.config['UPLOAD_FOLDER'], 'template.xlsx')
        wb.save(template_path)
        
        logger.info("Excel模板创建成功")
        
        # 直接返回文件下载
        return send_file(
            template_path,
            as_attachment=True,
            download_name='学生名单模板.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"模板下载失败: {e}")
        return jsonify({'success': False, 'message': '模板下载失败'}), 500

@app.route('/api/download/template', methods=['GET'])
def download_template_file():
    """下载模板文件"""
    try:
        template_path = os.path.join(app.config['UPLOAD_FOLDER'], 'template.xlsx')
        if os.path.exists(template_path):
            from flask import send_file
            return send_file(template_path, as_attachment=True, download_name='学生名单模板.xlsx')
        else:
            return jsonify({'success': False, 'message': '模板文件不存在'}), 404
    except Exception as e:
        logger.error(f"模板文件下载失败: {e}")
        return jsonify({'success': False, 'message': '文件下载失败'}), 500

@app.route('/lawSchUpl/api/upload/excel', methods=['POST'])
@permission_required
def upload_excel():
    """上传Excel文件"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': '没有选择文件'})
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'message': '没有选择文件'})
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'message': '只支持Excel文件格式'})
        
        # 保存文件
        filename = secure_filename(file.filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)
        
        # 读取Excel文件
        try:
            # 先读取第一行检查是否是提示信息
            df_check = pd.read_excel(file_path, nrows=1, header=None)
            first_cell = str(df_check.iloc[0, 0]) if not df_check.empty else ''
            
            # 如果第一行包含"重要提示"或"示例数据"，则从第2行开始读取（跳过提示行）
            skip_rows = 1 if ('重要提示' in first_cell or '示例数据' in first_cell or '⚠️' in first_cell) else 0
            
            # 指定某些列为字符串类型，避免数字格式问题
            # 手机号、学号、证件号码等应该作为字符串读取
            df = pd.read_excel(file_path, skiprows=skip_rows, dtype={
                '学号': str,
                '手机号码': str,
                '证件号码': str,
                '证件类型': str,
                '国家地区': str,
                '港澳台侨': str
            })
            
            logger.info(f"Excel文件读取成功，跳过行数: {skip_rows}，列名: {list(df.columns)}")
            logger.info(f"Excel文件数据行数: {len(df)}")
            if len(df) > 0:
                logger.info(f"第一行数据: {df.iloc[0].to_dict()}")
        except Exception as e:
            os.remove(file_path)
            return jsonify({'success': False, 'message': f'Excel文件读取失败: {str(e)}'})
        
        # 检查是否有数据
        if len(df) == 0:
            os.remove(file_path)
            return jsonify({
                'success': False, 
                'message': '文件中没有数据，请确保已删除示例数据行并填入真实学生信息'
            })
        
        # 数据验证
        validation_errors = []
        student_ids = set()
        
        for index, row in df.iterrows():
            # 计算Excel中的实际行号
            # 如果跳过了提示行，实际行号需要加上跳过的行数
            row_number = index + 2 + skip_rows  # Excel行号（1是提示行或表头，2是表头或数据）
            row_errors = validate_student_data(row, row_number)
            
            if row_errors:
                for error in row_errors:
                    validation_errors.append(f'第{row_number}行: {error}')
            
            # 检查学号重复（Excel内部）
            student_id = str(row.get('学号', '')).strip() if pd.notna(row.get('学号')) else ''
            name = str(row.get('姓名', '')).strip() if pd.notna(row.get('姓名')) else ''
            
            # 检查是否是模板示例数据
            if student_id == '025190210001' and name == '张三':
                validation_errors.append(
                    f'第{row_number}行: 检测到模板示例数据，请删除示例数据行后再上传真实学生信息'
                )
            
            if student_id:
                if student_id in student_ids:
                    validation_errors.append(f'第{row_number}行: Excel内学号重复 - 学号: {student_id}, 姓名: {name}')
                else:
                    student_ids.add(student_id)
                
                # 检查数据库中是否已存在相同学号和姓名的记录
                existing_record = LawSchoolUploadList.query.filter_by(
                    student_id=student_id,
                    name=name
                ).first()
                
                if existing_record:
                    validation_errors.append(
                        f'第{row_number}行: 数据库中已存在相同记录 - '
                        f'学号: {student_id}, 姓名: {name}, '
                        f'上传时间: {existing_record.create_time.strftime("%Y-%m-%d %H:%M:%S")}, '
                        f'上传人: {existing_record.uploader_id}'
                    )
        
        if validation_errors:
            os.remove(file_path)
            error_count = len(validation_errors)
            
            # 提取第一个错误的详细信息
            first_error = validation_errors[0]
            # 解析第一个错误，提取行号和字段信息
            first_error_highlight = first_error
            
            # 构建更友好的错误摘要，突出显示第一个错误
            error_summary = f'数据验证失败，未保存任何数据！{first_error_highlight}'
            
            # 限制返回的错误数量，避免响应过大
            max_errors = 50
            if error_count > max_errors:
                displayed_errors = validation_errors[:max_errors]
                displayed_errors.append(f'... 还有 {error_count - max_errors} 个错误未显示')
            else:
                displayed_errors = validation_errors
            
            return jsonify({
                'success': False,
                'message': error_summary,
                'errors': displayed_errors,
                'errorCount': error_count,
                'firstError': first_error  # 单独返回第一个错误，便于前端特殊显示
            })
        
        # 所有数据验证通过，开始批量入库
        uploader_id = session['user_id']
        success_count = 0
        inserted_records = []  # 记录插入的记录，用于跟踪
        
        try:
            # 批量插入数据 - 在try块中，任何异常都会触发回滚
            for index, row in df.iterrows():
                student_id = str(row.get('学号', '')).strip()
                
                # 调试日志：记录第一行数据的详细信息
                if index == 0:
                    logger.info(f"处理第一行数据，学号: {student_id}")
                    logger.info(f"民族: {row.get('民族')}, 国家地区: {row.get('国家地区')}")
                    logger.info(f"证件类型: {row.get('证件类型')}, 证件号码: {row.get('证件号码')}")
                    logger.info(f"入学年月: {row.get('入学年月')}, 培养层次: {row.get('培养层次')}")
                    logger.info(f"政治面貌: {row.get('政治面貌')}, 港澳台侨: {row.get('港澳台侨')}")
                
                # 创建记录（z308_id暂时为空）
                record = LawSchoolUploadList(
                    student_id=student_id,
                    name=str(row.get('姓名', '')).strip(),
                    gender=str(row.get('性别', '')).strip() if pd.notna(row.get('性别')) else None,
                    nationality=str(row.get('民族', '')).strip() if pd.notna(row.get('民族')) else None,
                    political_status=str(row.get('政治面貌', '')).strip() if pd.notna(row.get('政治面貌')) else None,
                    id_type=str(row.get('证件类型', '')).strip() if pd.notna(row.get('证件类型')) else None,
                    id_number=str(row.get('证件号码', '')).strip() if pd.notna(row.get('证件号码')) else None,
                    country=str(row.get('国家地区', '')).strip() if pd.notna(row.get('国家地区')) else None,
                    hong_kong_macao_taiwan=str(row.get('港澳台侨', '')).strip() if pd.notna(row.get('港澳台侨')) else None,
                    campus_email=str(row.get('校内电子邮箱', '')).strip() if pd.notna(row.get('校内电子邮箱')) else None,
                    personal_email=str(row.get('电子邮箱', '')).strip() if pd.notna(row.get('电子邮箱')) else None,
                    phone=clean_phone_number(row.get('手机号码')),
                    major=str(row.get('专业', '')).strip() if pd.notna(row.get('专业')) else None,
                    department=str(row.get('院系', '')).strip() if pd.notna(row.get('院系')) else None,
                    study_mode=str(row.get('学习方式', '')).strip() if pd.notna(row.get('学习方式')) else None,
                    grade=str(row.get('年级', '')).strip() if pd.notna(row.get('年级')) else None,
                    enrollment_date=pd.to_datetime(row.get('入学年月'), errors='coerce').date() if pd.notna(row.get('入学年月')) else None,
                    student_type=str(row.get('培养层次', '')).strip() if pd.notna(row.get('培养层次')) else None,
                    degree_type=str(row.get('学位类型', '')).strip() if pd.notna(row.get('学位类型')) else None,
                    special_plan=str(row.get('专项计划', '')).strip() if pd.notna(row.get('专项计划')) else None,
                    degree_level=str(row.get('培养层次', '')).strip() if pd.notna(row.get('培养层次')) else None,
                    expected_graduation_date=pd.to_datetime(row.get('预计毕业时间'), errors='coerce').date() if pd.notna(row.get('预计毕业时间')) else None,
                    enrollment_mode=str(row.get('入学方式名称', '')).strip() if pd.notna(row.get('入学方式名称')) else None,
                    student_category=str(row.get('学生类别', '')).strip() if pd.notna(row.get('学生类别')) else None,
                    supervisor=str(row.get('导师姓名', '')).strip() if pd.notna(row.get('导师姓名')) else None,
                    class_name=str(row.get('班级', '')).strip() if pd.notna(row.get('班级')) else None,
                    registration_status=str(row.get('注册状态', '')).strip() if pd.notna(row.get('注册状态')) else None,
                    z308_id=None,  # 初始为空，后续通过同步接口查询
                    sync_z308_id=0,  # 0-未查询z308_id
                    uploader_id=uploader_id
                )
                
                db.session.add(record)
                inserted_records.append(record)
                success_count += 1
            
            # 提交事务 - 只有所有记录都成功添加后才提交
            db.session.commit()
            logger.info(f"✅ 成功提交 {success_count} 条记录到数据库")
            
        except Exception as e:
            # 任何插入过程中的错误都会回滚整个事务
            logger.error(f"❌ 数据插入失败，回滚所有 {success_count} 条记录: {e}")
            db.session.rollback()
            os.remove(file_path)
            
            return jsonify({
                'success': False,
                'message': f'数据保存失败，所有数据已回滚，未保存任何记录',
                'errors': [f'数据库错误：{str(e)}'],
                'errorCount': 1,
                'firstError': f'数据库错误：{str(e)[:200]}'
            }), 500
        
        # 获取刚插入的记录ID，准备异步查询 z308_id
        student_records = []
        for record in inserted_records:
            student_records.append((record.id, record.student_id, record.name))
        
        # 创建异步任务
        task_id = str(uuid.uuid4())
        thread = threading.Thread(
            target=async_query_z308_ids,
            args=(task_id, student_records)
        )
        thread.daemon = True
        thread.start()
        
        # 删除临时文件
        os.remove(file_path)
        
        # 返回结果
        result = {
            'success': True,
            'message': f'上传成功！共导入 {success_count} 条记录，正在后台查询并同步Aleph账号...',
            'successCount': success_count,
            'failedCount': 0,  # 严格模式下失败数量为0
            'isProcessing': True,  # 标识正在处理中
            'taskId': task_id  # 返回任务ID，用于查询进度
        }
        
        return jsonify(result)
        
    except Exception as e:
        logger.error(f"文件上传失败: {e}")
        db.session.rollback()
        
        # 友好的错误提示
        error_message = str(e)
        
        # 检查是否是数据库约束错误
        if 'phone' in error_message and 'NULL' in error_message:
            return jsonify({
                'success': False,
                'message': '数据库配置错误：手机号字段需要允许为空',
                'errors': [
                    '数据库中的 phone 字段仍然设置为 NOT NULL，但系统已将手机号改为非必填项。',
                    '请联系数据库管理员执行 update_phone_column.sql 脚本来修改表结构。',
                    '详细说明请查看 UPDATE_DATABASE_INSTRUCTIONS.md 文件。'
                ],
                'errorCount': 3,
                'firstError': '数据库配置错误：phone 字段不允许 NULL 值，需要执行数据库更新脚本'
            }), 500
        elif 'IntegrityError' in error_message or 'UNIQUE' in error_message:
            return jsonify({
                'success': False,
                'message': '数据完整性错误：可能存在重复的学号',
                'errors': [f'数据库错误：{error_message}'],
                'errorCount': 1,
                'firstError': '数据库完整性约束冲突，请检查是否有重复的学号'
            }), 500
        else:
            return jsonify({
                'success': False,
                'message': f'文件上传失败：{error_message[:100]}',
                'errors': [f'系统错误：{error_message}'],
                'errorCount': 1,
                'firstError': f'系统错误：{error_message[:200]}'
            }), 500

@app.route('/lawSchUpl/api/task/progress/<task_id>', methods=['GET'])
@permission_required
def get_task_progress(task_id):
    """获取任务进度"""
    try:
        with task_lock:
            if task_id not in task_progress:
                return jsonify({
                    'success': False,
                    'message': '任务不存在或已过期'
                }), 404
            
            progress = task_progress[task_id].copy()
        
        return jsonify({
            'success': True,
            'data': progress
        })
    except Exception as e:
        logger.error(f"获取任务进度失败: {e}")
        return jsonify({
            'success': False,
            'message': '获取任务进度失败'
        }), 500

@app.route('/lawSchUpl/api/statistics', methods=['GET'])
@permission_required
def get_statistics():
    """获取统计数据"""
    try:
        from datetime import datetime, date
        
        # 总学生数
        total_count = LawSchoolUploadList.query.count()
        
        # 有Aleph账号的学生数
        has_aleph_count = LawSchoolUploadList.query.filter(
            LawSchoolUploadList.z308_id.isnot(None),
            LawSchoolUploadList.z308_id != ''
        ).count()
        
        # 无Aleph账号的学生数
        no_aleph_count = LawSchoolUploadList.query.filter(
            (LawSchoolUploadList.z308_id.is_(None)) | (LawSchoolUploadList.z308_id == '')
        ).count()
        
        # 未同步的学生数
        not_synced_count = LawSchoolUploadList.query.filter(
            LawSchoolUploadList.sync_z308_id == 0
        ).count()
        
        # 已同步的学生数
        synced_count = LawSchoolUploadList.query.filter(
            LawSchoolUploadList.sync_z308_id == 1
        ).count()
        
        # 同步失败的学生数
        sync_failed_count = LawSchoolUploadList.query.filter(
            LawSchoolUploadList.sync_z308_id == 2
        ).count()
        
        # 今日导入的学生数
        today = date.today()
        today_count = LawSchoolUploadList.query.filter(
            db.func.cast(LawSchoolUploadList.create_time, db.Date) == today
        ).count()
        
        # 本年导入的学生数
        this_year = today.year
        this_year_count = LawSchoolUploadList.query.filter(
            db.func.year(LawSchoolUploadList.create_time) == this_year
        ).count()
        
        return jsonify({
            'success': True,
            'data': {
                'totalCount': total_count,
                'hasAlephCount': has_aleph_count,
                'noAlephCount': no_aleph_count,
                'notSyncedCount': not_synced_count,
                'syncedCount': synced_count,
                'syncFailedCount': sync_failed_count,
                'todayCount': today_count,
                'thisYearCount': this_year_count
            }
        })
    except Exception as e:
        logger.error(f"获取统计数据失败: {str(e)}")
        return jsonify({'success': False, 'message': '获取统计数据失败'}), 500

@app.route('/lawSchUpl/api/students', methods=['GET'])
@permission_required
def get_students():
    """获取学生列表"""
    try:
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 20))
        uploader_id = request.args.get('uploader_id', '')
        student_id = request.args.get('student_id', '')
        name = request.args.get('name', '')
        sync_z308_id = request.args.get('sync_z308_id', '')
        
        # 构建查询
        query = LawSchoolUploadList.query
        
        if uploader_id:
            query = query.filter(LawSchoolUploadList.uploader_id.like(f'%{uploader_id}%'))
        if student_id:
            query = query.filter(LawSchoolUploadList.student_id.like(f'%{student_id}%'))
        if name:
            query = query.filter(LawSchoolUploadList.name.like(f'%{name}%'))
        if sync_z308_id:
            query = query.filter(LawSchoolUploadList.sync_z308_id == int(sync_z308_id))
        
        # 按创建时间倒序排列
        query = query.order_by(LawSchoolUploadList.create_time.desc())
        
        # 分页
        pagination = query.paginate(
            page=page, per_page=per_page, error_out=False
        )
        
        students = []
        for student in pagination.items:
            students.append({
                'id': student.id,
                'student_id': student.student_id,
                'name': student.name,
                'gender': student.gender,
                'nationality': student.nationality,
                'political_status': student.political_status,
                'id_type': student.id_type,
                'id_number': student.id_number,
                'country': student.country,
                'hong_kong_macao_taiwan': student.hong_kong_macao_taiwan,
                'campus_email': student.campus_email,
                'personal_email': student.personal_email,
                'phone': student.phone,
                'major': student.major,
                'department': student.department,
                'study_mode': student.study_mode,
                'grade': student.grade,
                'enrollment_date': student.enrollment_date.strftime('%Y-%m-%d') if student.enrollment_date else None,
                'student_type': student.student_type,
                'degree_type': student.degree_type,
                'special_plan': student.special_plan,
                'degree_level': student.degree_level,
                'expected_graduation_date': student.expected_graduation_date.strftime('%Y-%m-%d') if student.expected_graduation_date else None,
                'enrollment_mode': student.enrollment_mode,
                'student_category': student.student_category,
                'supervisor': student.supervisor,
                'class_name': student.class_name,
                'registration_status': student.registration_status,
                'z308_id': student.z308_id,
                'sync_z308_id': student.sync_z308_id,
                'uploader_id': student.uploader_id,
                'create_time': format_datetime(student.create_time),
                'update_time': format_datetime(student.update_time)
            })
        
        return jsonify({
            'success': True,
            'data': students,
            'pagination': {
                'page': page,
                'per_page': per_page,
                'total': pagination.total,
                'pages': pagination.pages
            }
        })
        
    except Exception as e:
        logger.error(f"获取学生列表失败: {e}")
        return jsonify({'success': False, 'message': '获取数据失败'}), 500

@app.route('/lawSchUpl/api/students/export', methods=['GET'])
@permission_required
def export_students():
    """导出学生数据"""
    try:
        uploader_id = request.args.get('uploader_id', '')
        student_id = request.args.get('student_id', '')
        name = request.args.get('name', '')
        sync_z308_id = request.args.get('sync_z308_id', '')
        
        # 构建查询
        query = LawSchoolUploadList.query
        
        if uploader_id:
            query = query.filter(LawSchoolUploadList.uploader_id.like(f'%{uploader_id}%'))
        if student_id:
            query = query.filter(LawSchoolUploadList.student_id.like(f'%{student_id}%'))
        if name:
            query = query.filter(LawSchoolUploadList.name.like(f'%{name}%'))
        if sync_z308_id:
            query = query.filter(LawSchoolUploadList.sync_z308_id == int(sync_z308_id))
        
        students = query.order_by(LawSchoolUploadList.create_time.desc()).all()
        
        # 转换为DataFrame
        data = []
        for student in students:
            # 确定同步状态文本
            if student.z308_id:
                if student.sync_z308_id == 1:
                    sync_status = '已同步'
                elif student.sync_z308_id == 2:
                    sync_status = '同步失败'
                else:
                    sync_status = '未同步'
            else:
                sync_status = '无Aleph账号'
            
            data.append({
                '学号': student.student_id,
                '姓名': student.name,
                '性别': student.gender,
                '民族': student.nationality,
                '政治面貌': student.political_status,
                '证件类型': student.id_type,
                '证件号码': student.id_number,
                '国家地区': student.country,
                '港澳台': student.hong_kong_macao_taiwan,
                '校内邮箱': student.campus_email,
                '个人邮箱': student.personal_email,
                '手机号码': student.phone,
                '专业': student.major,
                '院系': student.department,
                '学习方式': student.study_mode,
                '年级': student.grade,
                '入学日期': student.enrollment_date.strftime('%Y-%m-%d') if student.enrollment_date else '',
                '学生类型': student.student_type,
                '学位类型': student.degree_type,
                '专项计划': student.special_plan,
                '学位层次': student.degree_level,
                '预计毕业时间': student.expected_graduation_date.strftime('%Y-%m-%d') if student.expected_graduation_date else '',
                '入学方式': student.enrollment_mode,
                '学生类别': student.student_category,
                '导师姓名': student.supervisor,
                '班级': student.class_name,
                '注册状态': student.registration_status,
                'Aleph账号': student.z308_id if student.z308_id else '',
                '是否有Aleph账号': '有账号' if student.z308_id else '无账号',
                '同步状态': sync_status,
                '上传人学工号': student.uploader_id,
                '上传时间': format_datetime(student.create_time),
                '更新时间': format_datetime(student.update_time)
            })
        
        df = pd.DataFrame(data)
        
        # 保存为Excel文件
        export_path = os.path.join(app.config['UPLOAD_FOLDER'], f'export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
        df.to_excel(export_path, index=False)
        
        return jsonify({
            'success': True,
            'downloadUrl': f'/api/download/export?file={os.path.basename(export_path)}',
            'message': '导出成功'
        })
        
    except Exception as e:
        logger.error(f"导出数据失败: {e}")
        return jsonify({'success': False, 'message': '导出失败'}), 500

@app.route('/lawSchUpl/api/download/export', methods=['GET'])
def download_export_file():
    """下载导出文件"""
    try:
        filename = request.args.get('file', '')
        if not filename:
            return jsonify({'success': False, 'message': '文件名不能为空'}), 400
        
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        if os.path.exists(file_path):
            return send_file(file_path, as_attachment=True, download_name=filename)
        else:
            return jsonify({'success': False, 'message': '文件不存在'}), 404
    except Exception as e:
        logger.error(f"导出文件下载失败: {e}")
        return jsonify({'success': False, 'message': '文件下载失败'}), 500

@app.route('/lawSchUpl/api/upload/status', methods=['GET'])
@permission_required
def get_upload_status():
    """获取上传处理状态"""
    try:
        uploader_id = session['user_id']
        
        # 查询该上传人的数据统计
        total_count = LawSchoolUploadList.query.filter_by(uploader_id=uploader_id).count()
        has_z308_count = LawSchoolUploadList.query.filter(
            LawSchoolUploadList.uploader_id == uploader_id,
            LawSchoolUploadList.z308_id.isnot(None)
        ).count()
        no_z308_count = LawSchoolUploadList.query.filter(
            LawSchoolUploadList.uploader_id == uploader_id,
            LawSchoolUploadList.z308_id.is_(None),
            LawSchoolUploadList.sync_z308_id == 1
        ).count()
        not_queried_count = LawSchoolUploadList.query.filter_by(
            uploader_id=uploader_id, 
            sync_z308_id=0
        ).count()
        
        # 计算处理进度
        processed_count = has_z308_count + no_z308_count
        if total_count > 0:
            progress = (processed_count / total_count) * 100
        else:
            progress = 100
        
        # 判断是否还在处理中
        is_processing = not_queried_count > 0
        
        return jsonify({
            'success': True,
            'data': {
                'totalCount': total_count,
                'hasZ308Count': has_z308_count,
                'noZ308Count': no_z308_count,
                'notQueriedCount': not_queried_count,
                'progress': round(progress, 2),
                'isProcessing': is_processing,
                'message': f'共{total_count}条记录，有aleph账号{has_z308_count}条，无aleph账号{no_z308_count}条，未查询{not_queried_count}条'
            }
        })
        
    except Exception as e:
        logger.error(f"获取上传状态失败: {e}")
        return jsonify({'success': False, 'message': '获取状态失败'}), 500

@app.route('/lawSchUpl/api/students/<int:student_id>', methods=['PUT'])
@permission_required
def update_student(student_id):
    """更新单条学生数据"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': '请求数据不能为空'}), 400
        
        # 查找学生记录
        student = LawSchoolUploadList.query.get(student_id)
        if not student:
            return jsonify({'success': False, 'message': '学生记录不存在'}), 404
        
        # 检查权限：只能更新自己上传的数据
        if student.uploader_id != session['user_id']:
            return jsonify({'success': False, 'message': '无权限更新此学生数据'}), 403
        
        # 更新字段（只更新允许修改的字段）
        allowed_fields = [
            'name', 'gender', 'nationality', 'political_status', 'id_type', 'id_number',
            'country', 'hong_kong_macao_taiwan', 'campus_email', 'personal_email', 'phone',
            'major', 'department', 'study_mode', 'grade', 'enrollment_date', 'student_type',
            'degree_type', 'special_plan', 'degree_level', 'expected_graduation_date',
            'enrollment_mode', 'student_category', 'supervisor', 'class_name', 'registration_status'
        ]
        
        updated_fields = []
        for field in allowed_fields:
            if field in data:
                old_value = getattr(student, field)
                new_value = data[field]
                if old_value != new_value:
                    setattr(student, field, new_value)
                    updated_fields.append(field)
        
        # 如果学号发生变化，需要重新查询z308_id
        if 'student_id' in data and data['student_id'] != student.student_id:
            old_student_id = student.student_id
            student.student_id = data['student_id']
            updated_fields.append('student_id')
            
            # 重新查询z308_id
            try:
                z308_id = call_z308_id_api(student.student_id)
                if z308_id:
                    student.z308_id = z308_id
                    # 查询到z308_id后，sync_z308_id重置为0（未同步）
                    student.sync_z308_id = 0
                else:
                    student.z308_id = None
                    # 没有z308_id，sync_z308_id保持为0
                    student.sync_z308_id = 0
                updated_fields.append('z308_id')
                updated_fields.append('sync_z308_id')
            except Exception as e:
                logger.error(f"更新学号后查询z308_id失败: {e}")
                # 查询失败，sync_z308_id保持为0
        
        # 更新update_time
        student.update_time = get_local_time()
        updated_fields.append('update_time')
        
        # 提交事务
        db.session.commit()
        
        logger.info(f"学生数据更新成功: ID={student_id}, 更新字段={updated_fields}")
        
        return jsonify({
            'success': True,
            'message': '数据修改成功',
            'data': {
                'id': student.id,
                'student_id': student.student_id,
                'name': student.name,
                'updated_fields': updated_fields
            }
        })
        
    except Exception as e:
        logger.error(f"更新学生数据失败: {e}")
        db.session.rollback()
        return jsonify({'success': False, 'message': f'更新失败: {e}'}), 500

@app.route('/lawSchUpl/api/students/<int:student_id>/sync-aleph', methods=['POST'])
@permission_required
def sync_single_student_aleph(student_id):
    """
    单条学生记录同步Aleph账号
    1. 如果没有z308_id，先查询z308_id
    2. 如果有z308_id，调用同步接口
    3. 根据同步结果更新sync_z308_id字段
    """
    try:
        # 查找学生记录
        student = LawSchoolUploadList.query.get(student_id)
        if not student:
            return jsonify({'success': False, 'message': '学生记录不存在'}), 404
        
        logger.info(f"开始同步学生Aleph账号: ID={student_id}, 学号={student.student_id}, 姓名={student.name}")
        
        # 步骤1：如果没有z308_id，先查询
        if not student.z308_id:
            logger.info(f"学生 {student.student_id} 没有z308_id，先查询...")
            try:
                z308_id = call_z308_id_api(student.student_id)
                if z308_id:
                    student.z308_id = z308_id
                    logger.info(f"查询到z308_id: {z308_id}")
                else:
                    # 没有查询到z308_id，说明没有aleph账号
                    logger.warning(f"学生 {student.student_id} 没有Aleph账号")
                    student.update_time = get_local_time()
                    db.session.commit()
                    return jsonify({
                        'success': False,
                        'message': '该学生没有Aleph账号，无法同步'
                    })
            except Exception as e:
                logger.error(f"查询z308_id失败: {e}")
                return jsonify({
                    'success': False,
                    'message': f'查询Aleph账号失败: {str(e)}'
                }), 500
        
        # 步骤2：调用同步接口
        logger.info(f"开始调用同步接口，z308_id={student.z308_id}")
        student_data = {
            'student_id': student.student_id,
            'name': student.name,
            'gender': student.gender,
            'nationality': student.nationality,
            'political_status': student.political_status,
            'id_type': student.id_type,
            'id_number': student.id_number,
            'country': student.country,
            'hong_kong_macao_taiwan': student.hong_kong_macao_taiwan,
            'campus_email': student.campus_email,
            'personal_email': student.personal_email,
            'phone': student.phone,
            'major': student.major,
            'department': student.department,
            'study_mode': student.study_mode,
            'grade': student.grade,
            'enrollment_date': student.enrollment_date.strftime('%Y-%m-%d') if student.enrollment_date else None,
            'student_type': student.student_type,
            'degree_type': student.degree_type,
            'special_plan': student.special_plan,
            'degree_level': student.degree_level,
            'expected_graduation_date': student.expected_graduation_date.strftime('%Y-%m-%d') if student.expected_graduation_date else None,
            'enrollment_mode': student.enrollment_mode,
            'student_category': student.student_category,
            'supervisor': student.supervisor,
            'class_name': student.class_name,
            'registration_status': student.registration_status
        }
        
        success, message = call_sync_z308_id_api(student.z308_id, student_data)
        
        # 步骤3：根据同步结果更新sync_z308_id
        if success:
            student.sync_z308_id = 1  # 同步成功
            logger.info(f"同步成功: {message}")
        else:
            student.sync_z308_id = 2  # 同步失败
            logger.error(f"同步失败: {message}")
        
        student.update_time = get_local_time()
        db.session.commit()
        
        return jsonify({
            'success': success,
            'message': message,
            'data': {
                'id': student.id,
                'student_id': student.student_id,
                'name': student.name,
                'z308_id': student.z308_id,
                'sync_z308_id': student.sync_z308_id
            }
        })
        
    except Exception as e:
        logger.error(f"同步Aleph账号失败: {e}")
        db.session.rollback()
        return jsonify({'success': False, 'message': f'同步失败: {str(e)}'}), 500

@app.route('/lawSchUpl/api/trigger/z308-update', methods=['POST'])
@permission_required
def trigger_z308_update():
    """
    手动触发批量同步Aleph账号（支持筛选条件）
    处理逻辑：
    1. 查询所有sync_z308_id != 1的数据
    2. 对于没有z308_id的，先查询z308_id
    3. 对于有z308_id的，直接调用同步接口
    4. 根据同步结果更新sync_z308_id字段
    """
    try:
        # 获取筛选参数
        data = request.get_json() or {}
        uploader_id = data.get('uploader_id', '')
        student_id = data.get('student_id', '')
        name = data.get('name', '')
        sync_z308_id = data.get('sync_z308_id', '')
        
        # 构建查询条件：sync_z308_id != 1 的学生（未同步或同步失败）
        query = LawSchoolUploadList.query.filter(
            LawSchoolUploadList.sync_z308_id != 1
        )
        
        # 应用筛选条件
        if uploader_id:
            query = query.filter(LawSchoolUploadList.uploader_id.like(f'%{uploader_id}%'))
        if student_id:
            query = query.filter(LawSchoolUploadList.student_id.like(f'%{student_id}%'))
        if name:
            query = query.filter(LawSchoolUploadList.name.like(f'%{name}%'))
        if sync_z308_id != '':  # 允许筛选特定状态
            query = query.filter(LawSchoolUploadList.sync_z308_id == int(sync_z308_id))
        
        students = query.all()
        
        if not students:
            return jsonify({
                'success': True,
                'message': '没有需要同步的学生数据',
                'data': {
                    'totalCount': 0,
                    'syncedCount': 0,
                    'failedCount': 0,
                    'noAlephCount': 0
                }
            })
        
        logger.info(f"手动触发批量同步，共{len(students)}个学生需要处理")
        
        synced_count = 0  # 同步成功
        failed_count = 0  # 同步失败
        no_aleph_count = 0  # 没有aleph账号
        failed_students = []
        
        for student in students:
            try:
                # 步骤1：如果没有z308_id，先查询
                if not student.z308_id:
                    logger.info(f"学生 {student.student_id} 没有z308_id，先查询...")
                    z308_id = call_z308_id_api(student.student_id)
                    if z308_id:
                        student.z308_id = z308_id
                        logger.info(f"查询到z308_id: {z308_id}")
                    else:
                        # 没有查询到z308_id，说明没有aleph账号
                        logger.warning(f"学生 {student.student_id} 没有Aleph账号")
                        no_aleph_count += 1
                        student.update_time = get_local_time()
                        continue
                
                # 步骤2：调用同步接口
                logger.info(f"开始同步学生 {student.student_id}，z308_id={student.z308_id}")
                student_data = {
                    'student_id': student.student_id,
                    'name': student.name,
                    'gender': student.gender,
                    'nationality': student.nationality,
                    'political_status': student.political_status,
                    'id_type': student.id_type,
                    'id_number': student.id_number,
                    'country': student.country,
                    'hong_kong_macao_taiwan': student.hong_kong_macao_taiwan,
                    'campus_email': student.campus_email,
                    'personal_email': student.personal_email,
                    'phone': student.phone,
                    'major': student.major,
                    'department': student.department,
                    'study_mode': student.study_mode,
                    'grade': student.grade,
                    'enrollment_date': student.enrollment_date.strftime('%Y-%m-%d') if student.enrollment_date else None,
                    'student_type': student.student_type,
                    'degree_type': student.degree_type,
                    'special_plan': student.special_plan,
                    'degree_level': student.degree_level,
                    'expected_graduation_date': student.expected_graduation_date.strftime('%Y-%m-%d') if student.expected_graduation_date else None,
                    'enrollment_mode': student.enrollment_mode,
                    'student_category': student.student_category,
                    'supervisor': student.supervisor,
                    'class_name': student.class_name,
                    'registration_status': student.registration_status
                }
                
                success, message = call_sync_z308_id_api(student.z308_id, student_data)
                
                # 步骤3：根据同步结果更新sync_z308_id
                if success:
                    student.sync_z308_id = 1  # 同步成功
                    synced_count += 1
                    logger.info(f"同步成功: {message}")
                else:
                    student.sync_z308_id = 2  # 同步失败
                    failed_count += 1
                    logger.error(f"同步失败: {message}")
                    failed_students.append({
                        'student_id': student.student_id,
                        'name': student.name,
                        'error': message
                    })
                
                student.update_time = get_local_time()
                
            except Exception as e:
                logger.error(f"处理学生{student.student_id}时出错: {e}")
                student.sync_z308_id = 2  # 标记为同步失败
                failed_count += 1
                failed_students.append({
                    'student_id': student.student_id,
                    'name': student.name,
                    'error': str(e)
                })
        
        db.session.commit()
        
        logger.info(f"手动触发批量同步完成：成功{synced_count}个，失败{failed_count}个，无账号{no_aleph_count}个")
        
        return jsonify({
            'success': True,
            'message': f'同步完成！成功{synced_count}个，失败{failed_count}个，无Aleph账号{no_aleph_count}个',
            'data': {
                'totalCount': len(students),
                'syncedCount': synced_count,
                'failedCount': failed_count,
                'noAlephCount': no_aleph_count,
                'failedStudents': failed_students
            }
        })
        
    except Exception as e:
        logger.error(f"手动触发批量同步失败: {e}")
        db.session.rollback()
        return jsonify({'success': False, 'message': f'同步失败: {str(e)}'}), 500

# 异步查询z308_id函数
def update_z308_ids_for_uploader(uploader_id):
    """为指定上传人异步查询z308_id"""
    try:
        logger.info(f"开始为上传人{uploader_id}异步查询z308_id")
        
        # 查询该上传人未同步且有z308_id的学生（sync_z308_id=0 且 z308_id不为空）
        students = LawSchoolUploadList.query.filter(
            LawSchoolUploadList.uploader_id == uploader_id,
            LawSchoolUploadList.sync_z308_id == 0,
            LawSchoolUploadList.z308_id.isnot(None),
            LawSchoolUploadList.z308_id != ''
        ).all()
        
        updated_count = 0
        failed_count = 0
        
        for student in students:
            try:
                z308_id = call_z308_id_api(student.student_id)
                if z308_id:
                    # 查询到z308_id，表示有aleph账号
                    student.z308_id = z308_id
                    # 这里需要调用同步接口，暂时标记为已查询到z308_id
                    # 后续需要根据同步接口的返回结果来设置sync_z308_id
                    student.sync_z308_id = 1  # 标记为已查询到z308_id
                    student.update_time = get_local_time()
                    updated_count += 1
                else:
                    # 未查询到z308_id，表示无aleph账号
                    student.z308_id = None
                    student.sync_z308_id = 1  # 标记为已处理（无aleph账号）
                    student.update_time = get_local_time()
                    updated_count += 1
            except Exception as e:
                logger.error(f"更新学生{student.student_id}的z308_id失败: {e}")
                student.sync_z308_id = 2  # 标记为查询失败
                failed_count += 1
        
        db.session.commit()
        logger.info(f"异步查询完成：成功更新{updated_count}个，失败{failed_count}个学生的z308_id")
        
    except Exception as e:
        logger.error(f"异步查询z308_id失败: {e}")
        db.session.rollback()

# 定时任务：每天凌晨5点同步Aleph账号
def update_z308_ids():
    """
    定时任务：批量同步Aleph账号
    处理逻辑：
    1. 查询所有sync_z308_id != 1的数据（全表）
    2. 对于没有z308_id的，先查询z308_id
    3. 对于有z308_id的，直接调用同步接口
    4. 根据同步结果更新sync_z308_id字段
    """
    with app.app_context():
        try:
            logger.info("开始执行定时任务：批量同步Aleph账号")
            
            # 查询所有sync_z308_id != 1的学生（未同步或同步失败）
            students = LawSchoolUploadList.query.filter(
                LawSchoolUploadList.sync_z308_id != 1
            ).all()
            
            logger.info(f"定时任务：共{len(students)}个学生需要处理")
            
            synced_count = 0  # 同步成功
            failed_count = 0  # 同步失败
            no_aleph_count = 0  # 没有aleph账号
            
            for student in students:
                try:
                    # 步骤1：如果没有z308_id，先查询
                    if not student.z308_id:
                        logger.info(f"学生 {student.student_id} 没有z308_id，先查询...")
                        z308_id = call_z308_id_api(student.student_id)
                        if z308_id:
                            student.z308_id = z308_id
                            logger.info(f"查询到z308_id: {z308_id}")
                        else:
                            # 没有查询到z308_id，说明没有aleph账号
                            logger.warning(f"学生 {student.student_id} 没有Aleph账号")
                            no_aleph_count += 1
                            student.update_time = get_local_time()
                            continue
                    
                    # 步骤2：调用同步接口
                    logger.info(f"开始同步学生 {student.student_id}，z308_id={student.z308_id}")
                    student_data = {
                        'student_id': student.student_id,
                        'name': student.name,
                        'gender': student.gender,
                        'nationality': student.nationality,
                        'political_status': student.political_status,
                        'id_type': student.id_type,
                        'id_number': student.id_number,
                        'country': student.country,
                        'hong_kong_macao_taiwan': student.hong_kong_macao_taiwan,
                        'campus_email': student.campus_email,
                        'personal_email': student.personal_email,
                        'phone': student.phone,
                        'major': student.major,
                        'department': student.department,
                        'study_mode': student.study_mode,
                        'grade': student.grade,
                        'enrollment_date': student.enrollment_date.strftime('%Y-%m-%d') if student.enrollment_date else None,
                        'student_type': student.student_type,
                        'degree_type': student.degree_type,
                        'special_plan': student.special_plan,
                        'degree_level': student.degree_level,
                        'expected_graduation_date': student.expected_graduation_date.strftime('%Y-%m-%d') if student.expected_graduation_date else None,
                        'enrollment_mode': student.enrollment_mode,
                        'student_category': student.student_category,
                        'supervisor': student.supervisor,
                        'class_name': student.class_name,
                        'registration_status': student.registration_status
                    }
                    
                    success, message = call_sync_z308_id_api(student.z308_id, student_data)
                    
                    # 步骤3：根据同步结果更新sync_z308_id
                    if success:
                        student.sync_z308_id = 1  # 同步成功
                        synced_count += 1
                        logger.info(f"同步成功: {message}")
                    else:
                        student.sync_z308_id = 2  # 同步失败
                        failed_count += 1
                        logger.error(f"同步失败: {message}")
                    
                    student.update_time = get_local_time()
                    
                except Exception as e:
                    logger.error(f"处理学生{student.student_id}时出错: {e}")
                    student.sync_z308_id = 2  # 标记为同步失败
                    failed_count += 1
            
            db.session.commit()
            logger.info(f"定时任务完成：成功{synced_count}个，失败{failed_count}个，无账号{no_aleph_count}个")
            
        except Exception as e:
            logger.error(f"定时任务执行失败: {e}")
            db.session.rollback()

# 配置定时任务
scheduler = BackgroundScheduler()
scheduler.add_job(
    func=update_z308_ids,
    trigger="cron",
    hour=5,
    minute=0,
    id='update_z308_ids_job'
)
scheduler.add_job(
    func=clean_old_tasks,
    trigger="interval",
    hours=1,
    id='clean_old_tasks_job'
)
scheduler.start()

# 健康检查端点
@app.route('/lawSchUpl/api/health')
def health_check():
    """健康检查端点"""
    try:
        # 检查数据库连接
        from sqlalchemy import text
        db.session.execute(text('SELECT 1'))
        return jsonify({
            'status': 'healthy',
            'timestamp': get_local_time().isoformat(),
            'database': 'connected'
        }), 200
    except Exception as e:
        logger.error(f"Health check failed: {str(e)}")
        return jsonify({
            'status': 'unhealthy',
            'timestamp': get_local_time().isoformat(),
            'error': str(e)
        }), 500

# 关闭时停止定时任务
atexit.register(lambda: scheduler.shutdown())

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(debug=True, host='0.0.0.0', port=5000)
